'''Дан эксельный файл. На странице находится список сотрудников и их ЗП.
Надо создать еще одну страницу в этом файле и поместить туда отсортированный
список фамилий и их ЗП, а в последней строчке поместить слово ИТОГО: и сумму
всех ЗП. Сортировка по убыванию ЗП сотрудников.'''

import openpyxl

workbook = openpyxl.load_workbook('employees.xlsx')
sheet = workbook.active
employee_data = [(row[0].value, row[1].value) for row in sheet.iter_rows(values_only=True)]
sorted_employee_data = sorted(employee_data, key=lambda x: x[1], reverse=True)
new_sheet = workbook.create_sheet(title='Sorted Employees')

for idx, (name, salary) in enumerate(sorted_employee_data, start=1):
    new_sheet.cell(row=idx, column=1, value=name)
    new_sheet.cell(row=idx, column=2, value=salary)

total_salary = sum(salary for name, salary in employee_data)
new_sheet.cell(row=len(sorted_employee_data)+2, column=1, value='ИТОГО:')
new_sheet.cell(row=len(sorted_employee_data)+2, column=2, value=total_salary)

workbook.save('employees.xlsx')

