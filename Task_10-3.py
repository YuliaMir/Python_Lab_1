'''Дан исходный эксельный файл со списком людей и их ЗП.
Следует создать еще одну страницу со статистическими данными исходного списка.'''

import openpyxl
from statistics import median

workbook = openpyxl.load_workbook('employees.xlsx')
sheet = workbook['Sheet']
salaries = [row[1].value for row in sheet.iter_rows(values_only=True)][1:]


max_salary = max(salaries)
min_salary = min(salaries)
total_salary = sum(salaries)
mean_salary = total_salary / len(salaries)
median_salary = median(salaries)

new_sheet = workbook.create_sheet(title='Statistics')

new_sheet.cell(row=1, column=1, value='Maximum Value')
new_sheet.cell(row=1, column=2, value=max_salary)

new_sheet.cell(row=2, column=1, value='Minimum Value')
new_sheet.cell(row=2, column=2, value=min_salary)

new_sheet.cell(row=3, column=1, value='Sum')
new_sheet.cell(row=3, column=2, value=total_salary)

new_sheet.cell(row=4, column=1, value='Arithmetic Mean')
new_sheet.cell(row=4, column=2, value=mean_salary)

new_sheet.cell(row=5, column=1, value='Median')
new_sheet.cell(row=5, column=2, value=median_salary)

workbook.save('employees.xlsx')


